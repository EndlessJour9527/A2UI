# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Unit tests for Excel test-set parser."""

from __future__ import annotations

import openpyxl
from pathlib import Path
from genui_eval.excel_parser import parse_excel_test_set


def test_excel_parser_parses_valid_sheet(tmp_path: Path):
    excel_path = tmp_path / "test_set.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Starter Sheet"

    # Write headers
    sheet.append(
        [
            "Prompt",
            "Group",
            "Description",
            "Target",
            "Renderer",
            "Spec Version",
            "Catalog ID",
            "Catalog Profile ID",
        ]
    )
    # Row 1
    sheet.append(
        [
            "Render a card",
            "starter-group",
            "Card case",
            "Expect card",
            "ink",
            "0.9",
            "https://a2ui.org/catalog.json",
            "ink-a2ui-v0_9",
        ]
    )
    # Row 2
    sheet.append(
        [
            "Render a button",
            "",  # empty group -> should use sheet name slugified
            "Button case",
            "Expect button",
            "react",
            "0.8",
            "",
            "",
        ]
    )

    wb.save(str(excel_path))

    groups = parse_excel_test_set(excel_path)
    # We should have two groups: "starter-group" and "starter-sheet"
    assert len(groups) == 2

    # Map groups by ID for checking
    group_map = {g.group_id: g for g in groups}
    assert "starter-group" in group_map
    assert "starter-sheet" in group_map

    g1 = group_map["starter-group"]
    assert len(g1.cases) == 1
    c1 = g1.cases[0]
    assert c1.case_id == "case-1"
    assert c1.prompt == "Render a card"
    assert c1.renderer == "ink"
    assert c1.spec_version == "0.9"
    assert c1.catalog_id == "https://a2ui.org/catalog.json"
    assert c1.catalog_profile_id == "ink-a2ui-v0_9"

    g2 = group_map["starter-sheet"]
    assert len(g2.cases) == 1
    c2 = g2.cases[0]
    assert c2.case_id == "case-2"
    assert c2.prompt == "Render a button"
    assert c2.renderer == "react"
    assert c2.spec_version == "0.8"
    assert c2.catalog_id is None
    assert c2.catalog_profile_id is None


def test_excel_parser_multiple_sheets(tmp_path: Path):
    excel_path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1
    sheet1 = wb.active
    sheet1.title = "First Sheet"
    sheet1.append(["Prompt", "Target"])
    sheet1.append(["Prompt 1", "Target 1"])

    # Sheet 2
    sheet2 = wb.create_sheet(title="Second Sheet")
    sheet2.append(["Prompt", "Target"])
    sheet2.append(["Prompt 2", "Target 2"])

    wb.save(str(excel_path))

    groups = parse_excel_test_set(excel_path)
    assert len(groups) == 2

    group_map = {g.group_id: g for g in groups}
    assert "first-sheet" in group_map
    assert "second-sheet" in group_map

    assert len(group_map["first-sheet"].cases) == 1
    assert group_map["first-sheet"].cases[0].prompt == "Prompt 1"

    assert len(group_map["second-sheet"].cases) == 1
    assert group_map["second-sheet"].cases[0].prompt == "Prompt 2"


def test_excel_parser_parses_protocol_columns(tmp_path: Path):
    excel_path = tmp_path / "protocol_columns.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Protocol Cases"
    sheet.append(
        [
            "prompt",
            "target",
            "protocol_id",
            "protocol_profile_id",
            "protocol_version",
            "protocol_options",
            "catalog_profile_id",
        ]
    )
    sheet.append(
        [
            "Render an OpenUI card",
            '{"kind":"card"}',
            "openui",
            "openui-default-v1",
            "1",
            '{"mode":"skeleton"}',
            "",
        ]
    )
    sheet.append(
        [
            "Render an A2UI card",
            "Expect A2UI card",
            "a2ui",
            "a2ui-basic-v0_9",
            "0.9",
            "{}",
            "a2ui-basic-v0_9",
        ]
    )
    wb.save(str(excel_path))

    groups = parse_excel_test_set(excel_path)
    cases = groups[0].cases

    assert cases[0].protocol_id == "openui"
    assert cases[0].protocol_version == "1"
    assert cases[0].protocol_profile_id == "openui-default-v1"
    assert cases[0].protocol_options == {"mode": "skeleton"}
    assert cases[0].catalog_profile_id is None

    assert cases[1].protocol_id == "a2ui"
    assert cases[1].protocol_version == "0.9"
    assert cases[1].protocol_profile_id == "a2ui-basic-v0_9"
    assert cases[1].protocol_options["catalogProfileId"] == "a2ui-basic-v0_9"
    assert cases[1].catalog_profile_id == "a2ui-basic-v0_9"


def test_create_run_from_excel_script(tmp_path: Path):
    from unittest.mock import patch
    import sys
    import io
    import json
    from contextlib import redirect_stdout

    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "bin"))
    import create_run_from_excel

    excel_path = tmp_path / "test_set.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Sheet"
    sheet.append(["Prompt", "Target"])
    sheet.append(["Test Prompt", "Test Target"])
    wb.save(str(excel_path))

    studio_root = tmp_path / ".genui-eval-studio"

    test_args = [
        "create_run_from_excel.py",
        "--file", str(excel_path),
        "--model", "test-model",
        "--studio-root", str(studio_root)
    ]

    with patch.object(sys, "argv", test_args):
        f = io.StringIO()
        with redirect_stdout(f):
            create_run_from_excel.main()

        output = f.getvalue()
        result = json.loads(output)
        assert "runId" in result
        assert result["totalCases"] == 1

        run_id = result["runId"]
        run_dir = studio_root / "runs" / run_id
        assert (run_dir / "source" / "source.xlsx").exists()
        assert (run_dir / "manifest.json").exists()
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        assert manifest["artifacts"]["source_excel"] == "source/source.xlsx"


def test_excel_parser_parses_json_cases(tmp_path: Path):
    json_path = tmp_path / "test_set.json"
    import json
    data = [
        {
            "prompt": "Render a dynamic card",
            "group": "json-flat-group",
            "description": "Card description",
            "target": "Card target",
            "renderer": "react",
            "spec_version": "0.9"
        },
        {
            "prompt": "Render a dynamic button",
            "group": "json-flat-group",
            "target": "Button target",
            "renderer": "ink"
        }
    ]
    json_path.write_text(json.dumps(data), encoding="utf-8")

    groups = parse_excel_test_set(json_path)
    assert len(groups) == 1
    assert groups[0].group_id == "json-flat-group"
    assert len(groups[0].cases) == 2
    assert groups[0].cases[0].prompt == "Render a dynamic card"
    assert groups[0].cases[0].renderer == "react"
    assert groups[0].cases[1].prompt == "Render a dynamic button"
    assert groups[0].cases[1].renderer == "ink"


def test_excel_parser_parses_json_groups(tmp_path: Path):
    json_path = tmp_path / "test_groups.json"
    import json
    data = [
        {
            "group_id": "group-a",
            "label": "Group A",
            "cases": [
                {
                    "prompt": "Prompt A1",
                    "target": "Target A1"
                }
            ]
        },
        {
            "group_id": "group-b",
            "label": "Group B",
            "cases": [
                {
                    "prompt": "Prompt B1"
                }
            ]
        }
    ]
    json_path.write_text(json.dumps(data), encoding="utf-8")

    groups = parse_excel_test_set(json_path)
    assert len(groups) == 2
    group_map = {g.group_id: g for g in groups}
    assert "group-a" in group_map
    assert "group-b" in group_map
    assert group_map["group-a"].label == "Group A"
    assert group_map["group-a"].cases[0].prompt == "Prompt A1"
    assert group_map["group-b"].cases[0].prompt == "Prompt B1"


def test_create_run_from_json_script(tmp_path: Path):
    from unittest.mock import patch
    import sys
    import io
    import json
    from contextlib import redirect_stdout

    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "bin"))
    import create_run_from_excel

    json_path = tmp_path / "test_set.json"
    data = [
        {"prompt": "Test Prompt JSON", "target": "Test Target JSON"}
    ]
    json_path.write_text(json.dumps(data), encoding="utf-8")

    studio_root = tmp_path / ".genui-eval-studio"

    test_args = [
        "create_run_from_excel.py",
        "--file", str(json_path),
        "--model", "test-model-json",
        "--studio-root", str(studio_root)
    ]

    with patch.object(sys, "argv", test_args):
        f = io.StringIO()
        with redirect_stdout(f):
            create_run_from_excel.main()

        output = f.getvalue()
        result = json.loads(output)
        assert "runId" in result
        assert result["totalCases"] == 1

        run_id = result["runId"]
        run_dir = studio_root / "runs" / run_id
        assert (run_dir / "source" / "source.json").exists()
        assert (run_dir / "manifest.json").exists()
        manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
        assert manifest["artifacts"]["source_json"] == "source/source.json"


def test_excel_parser_custom_columns_in_protocol_options(tmp_path: Path):
    excel_path = tmp_path / "custom_cols.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Custom Cols"
    sheet.append(
        [
            "prompt",
            "group",
            "my_custom_option",
            "another_option",
        ]
    )
    sheet.append(
        [
            "Test prompt with custom options",
            "custom-group",
            "val1",
            "val2",
        ]
    )
    wb.save(str(excel_path))

    groups = parse_excel_test_set(excel_path)
    assert len(groups) == 1
    cases = groups[0].cases
    assert len(cases) == 1
    assert cases[0].protocol_options["my_custom_option"] == "val1"
    assert cases[0].protocol_options["another_option"] == "val2"


def test_json_parser_custom_keys_in_protocol_options(tmp_path: Path):
    json_path = tmp_path / "custom_keys.json"
    import json
    data = [
        {
            "prompt": "Test prompt json custom",
            "group": "custom-group",
            "custom_key_json": "json_val",
            "extra_field": 42
        }
    ]
    json_path.write_text(json.dumps(data), encoding="utf-8")

    groups = parse_excel_test_set(json_path)
    assert len(groups) == 1
    cases = groups[0].cases
    assert len(cases) == 1
    assert cases[0].protocol_options["custom_key_json"] == "json_val"
    assert cases[0].protocol_options["extra_field"] == 42


